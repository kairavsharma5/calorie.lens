import openpyxl

wb = openpyxl.load_workbook('Anuvaad_INDB_2024.11.xlsx', data_only=True)

print("Sheet names:", wb.sheetnames)

sheet = wb[wb.sheetnames[0]]
print("\nFirst row (headers):")
for cell in sheet[1]:
    print(cell.value)

print("\nSecond row (sample data):")
for cell in sheet[2]:
    print(cell.value)